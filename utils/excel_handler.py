#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from datetime import datetime

class ExcelHandler:
    # Column mapping
    COLUMN_MAPPING = {
        'nim': ['nim', 'NIM', 'no', 'nomor induk'],
        'nama': ['nama', 'NAMA', 'nama mahasiswa', 'full name'],
        'prodi': ['prodi', 'PRODI', 'program studi', 'jurusan', 'program'],
        'tanggal_lahir': ['tanggal lahir', 'tanggal_lahir', 'dob', 'birth date'],
        'alamat': ['alamat', 'ALAMAT', 'address']
    }
    
    @staticmethod
    def find_column_index(headers, field):
        """Cari index kolom berdasarkan field"""
        field_aliases = ExcelHandler.COLUMN_MAPPING.get(field, [])
        
        for idx, header in enumerate(headers):
            if header.strip().lower() in field_aliases:
                return idx
        return None
    
    @staticmethod
    def parse_excel(file):
        """Parse file Excel dan ekstrak data mahasiswa"""
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            if not ws or ws.max_row < 2:
                return None, 'File Excel kosong atau tidak memiliki data'
            
            # Baca header dari baris pertama
            headers = [str(cell.value).strip() for cell in ws[1]]
            
            # Cari index kolom yang diperlukan
            nim_idx = ExcelHandler.find_column_index(headers, 'nim')
            nama_idx = ExcelHandler.find_column_index(headers, 'nama')
            prodi_idx = ExcelHandler.find_column_index(headers, 'prodi')
            tanggal_lahir_idx = ExcelHandler.find_column_index(headers, 'tanggal_lahir')
            alamat_idx = ExcelHandler.find_column_index(headers, 'alamat')
            
            # Validasi kolom penting
            if nim_idx is None or nama_idx is None or prodi_idx is None:
                return None, 'File Excel harus memiliki kolom: NIM, Nama, dan Prodi'
            
            mahasiswa_list = []
            
            # Baca data dari baris ke-2 sampai akhir
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                nim = row[nim_idx].value
                nama = row[nama_idx].value
                prodi = row[prodi_idx].value
                tanggal_lahir = row[tanggal_lahir_idx].value if tanggal_lahir_idx is not None else None
                alamat = row[alamat_idx].value if alamat_idx is not None else None
                
                # Skip baris kosong
                if not nim or not nama or not prodi:
                    continue
                
                # Format tanggal jika ada
                tanggal_str = ''
                if tanggal_lahir:
                    try:
                        if isinstance(tanggal_lahir, datetime):
                            tanggal_str = tanggal_lahir.strftime('%Y-%m-%d')
                        else:
                            # Coba parse string
                            dt = datetime.strptime(str(tanggal_lahir), '%d/%m/%Y')
                            tanggal_str = dt.strftime('%Y-%m-%d')
                    except:
                        tanggal_str = str(tanggal_lahir)
                
                mahasiswa_list.append({
                    'nim': str(nim).strip(),
                    'nama': str(nama).strip(),
                    'prodi': str(prodi).strip(),
                    'tanggal_lahir': tanggal_str,
                    'alamat': str(alamat).strip() if alamat else ''
                })
            
            if not mahasiswa_list:
                return None, 'Tidak ada data mahasiswa ditemukan di file Excel'
            
            return mahasiswa_list, None
        
        except Exception as e:
            return None, f'Error membaca file Excel: {str(e)}'
